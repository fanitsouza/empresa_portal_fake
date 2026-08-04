from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from extracao_pdf import CAMPOS_OBRIGATORIOS


NOME_ABA = "Dados extraidos"
COLUNAS = [
    "Arquivo",
    *CAMPOS_OBRIGATORIOS,
    "Sucesso",
    "Campos ausentes",
    "Erro",
    "Atualizado em",
]


def atualizar_planilha_mestra(
    resultados: list[dict[str, Any]], caminho_planilha: str | Path
) -> tuple[Path, int, int]:
    """Cria ou atualiza a planilha, evitando duplicidade por CPF ou arquivo."""
    caminho = Path(caminho_planilha)
    caminho.parent.mkdir(parents=True, exist_ok=True)

    if caminho.exists():
        workbook = load_workbook(caminho)
        planilha = workbook[NOME_ABA] if NOME_ABA in workbook.sheetnames else workbook.active
        _validar_cabecalho(planilha)
    else:
        workbook = Workbook()
        planilha = workbook.active
        planilha.title = NOME_ABA
        planilha.append(COLUNAS)
        _formatar_planilha(planilha)

    linhas_existentes = _indexar_linhas(planilha)
    adicionados = 0
    atualizados = 0

    for resultado in resultados:
        valores = _resultado_para_linha(resultado)
        chave = _chave_resultado(resultado)
        numero_linha = linhas_existentes.get(chave)

        if numero_linha is None:
            planilha.append(valores)
            numero_linha = planilha.max_row
            linhas_existentes[chave] = numero_linha
            adicionados += 1
        else:
            for coluna, valor in enumerate(valores, start=1):
                planilha.cell(numero_linha, coluna, valor)
            atualizados += 1

    planilha.auto_filter.ref = planilha.dimensions
    planilha.freeze_panes = "A2"
    workbook.save(caminho)
    return caminho, adicionados, atualizados


def _resultado_para_linha(resultado: dict[str, Any]) -> list[Any]:
    dados = resultado.get("dados") or {}
    return [
        resultado.get("arquivo", ""),
        *(dados.get(campo, "") for campo in CAMPOS_OBRIGATORIOS),
        "Sim" if resultado.get("sucesso") else "Nao",
        ", ".join(resultado.get("campos_ausentes") or []),
        resultado.get("erro") or "",
        datetime.now().astimezone().strftime("%d/%m/%Y %H:%M:%S"),
    ]


def _chave_resultado(resultado: dict[str, Any]) -> str:
    cpf = (resultado.get("dados") or {}).get("CPF", "")
    cpf_normalizado = re.sub(r"\D", "", cpf)
    if cpf_normalizado:
        return f"cpf:{cpf_normalizado}"
    return f"arquivo:{Path(resultado.get('arquivo', '')).resolve()}".casefold()


def _indexar_linhas(planilha) -> dict[str, int]:
    indices = {nome: indice + 1 for indice, nome in enumerate(COLUNAS)}
    linhas: dict[str, int] = {}
    for numero_linha in range(2, planilha.max_row + 1):
        cpf = str(planilha.cell(numero_linha, indices["CPF"]).value or "")
        arquivo = str(planilha.cell(numero_linha, indices["Arquivo"]).value or "")
        cpf_normalizado = re.sub(r"\D", "", cpf)
        chave = (
            f"cpf:{cpf_normalizado}"
            if cpf_normalizado
            else f"arquivo:{Path(arquivo).resolve()}".casefold()
        )
        linhas[chave] = numero_linha
    return linhas


def _validar_cabecalho(planilha) -> None:
    cabecalho = [planilha.cell(1, coluna).value for coluna in range(1, len(COLUNAS) + 1)]
    if cabecalho != COLUNAS:
        raise ValueError(
            "A planilha existente nao possui o formato esperado e nao foi alterada."
        )


def _formatar_planilha(planilha) -> None:
    preenchimento = PatternFill("solid", fgColor="1F4E78")
    for celula in planilha[1]:
        celula.font = Font(color="FFFFFF", bold=True)
        celula.fill = preenchimento

    larguras = {
        "A": 42,
        "B": 28,
        "C": 18,
        "D": 30,
        "E": 20,
        "F": 16,
        "G": 42,
        "H": 12,
        "I": 30,
        "J": 45,
        "K": 22,
    }
    for coluna, largura in larguras.items():
        planilha.column_dimensions[coluna].width = largura
