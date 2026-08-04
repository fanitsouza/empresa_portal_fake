import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.append(str(Path(__file__).resolve().parents[1] / "source"))

from planilha_mestra import COLUNAS, atualizar_planilha_mestra  # noqa: E402


def _resultado(cpf="123.456.789-01", nome="Ana"):
    return {
        "arquivo": "documento.pdf",
        "sucesso": True,
        "dados": {
            "Nome": nome,
            "CPF": cpf,
            "E-mail": "ana@exemplo.com",
            "Telefone": "(92) 99999-9999",
            "Nascimento": "01/01/2000",
            "Endereco": "Rua A",
        },
        "campos_ausentes": [],
        "erro": None,
    }


def test_cria_planilha_mestra(tmp_path):
    caminho = tmp_path / "planilha_mestra.xlsx"

    _, adicionados, atualizados = atualizar_planilha_mestra([_resultado()], caminho)

    planilha = load_workbook(caminho).active
    assert [celula.value for celula in planilha[1]] == COLUNAS
    assert planilha.max_row == 2
    assert planilha["B2"].value == "Ana"
    assert (adicionados, atualizados) == (1, 0)


def test_atualiza_cpf_existente_sem_duplicar(tmp_path):
    caminho = tmp_path / "planilha_mestra.xlsx"
    atualizar_planilha_mestra([_resultado()], caminho)

    _, adicionados, atualizados = atualizar_planilha_mestra(
        [_resultado(nome="Ana Atualizada")], caminho
    )

    planilha = load_workbook(caminho).active
    assert planilha.max_row == 2
    assert planilha["B2"].value == "Ana Atualizada"
    assert (adicionados, atualizados) == (0, 1)


def test_adiciona_cpf_novo(tmp_path):
    caminho = tmp_path / "planilha_mestra.xlsx"
    atualizar_planilha_mestra([_resultado()], caminho)

    _, adicionados, atualizados = atualizar_planilha_mestra(
        [_resultado(cpf="987.654.321-00", nome="Bruno")], caminho
    )

    planilha = load_workbook(caminho).active
    assert planilha.max_row == 3
    assert planilha["B3"].value == "Bruno"
    assert (adicionados, atualizados) == (1, 0)
